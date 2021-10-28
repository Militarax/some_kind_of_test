import os
from os import listdir
from os.path import isfile, join
from openpyxl import load_workbook


def main():
	data_dir = os.path.join(os.getcwd(), 'data')
	xlsx_files = [os.path.join(data_dir, file) for file in listdir(data_dir) if isfile(join(data_dir, file)) and file.endswith('.xlsx')]
	
	for filename in xlsx_files:

		wb = load_workbook(filename=filename)
		sheet_names = wb.sheetnames

		for sheet_name in sheet_names:

			date = wb[sheet_name]['A3'].value
			week_start = date[date.find('From:') + len('From:'): date.find('To:')].strip()
			week_end = date[date.find('To:') + len('To:'):].strip()

			new_filename = week_start + '_' + week_end + '.csv'

			with open(os.path.join(data_dir, new_filename), 'w') as file:
				for row in range(7, 31):
					key = wb[sheet_name]['A' + str(row)].value 
					amount = wb[sheet_name]['C' + str(row)].value

					if not key:
						continue

					file.write(','.join([key.replace(',', ' ') , str(amount) + '\n']))
				
				file.write(','.join(['week_start', week_start + '\n']))			# just to be sure that we have start and end in the file
				file.write(','.join(['week_end', week_end + '\n']))



if __name__ == '__main__':
	main()
